"""Financial summaries and CSV/Excel export."""

from __future__ import annotations

import csv
import io
import statistics
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional
from zoneinfo import ZoneInfo

from pesa_logger.database import list_transactions, log_report_run


_REPORT_TZ = ZoneInfo("Africa/Nairobi")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _parse_ts(ts: Optional[str]) -> Optional[datetime]:
    if not ts:
        return None
    try:
        dt = datetime.fromisoformat(ts)
    except (ValueError, TypeError):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_REPORT_TZ)


def _group_by_period(transactions: list, period: str) -> Dict[str, list]:
    """Group transactions by 'weekly' or 'monthly' period key."""
    groups: Dict[str, list] = defaultdict(list)
    for tx in transactions:
        dt = _parse_ts(tx.get("timestamp"))
        if not dt:
            continue
        if period == "weekly":
            # ISO week: e.g. "2026-W07"
            key = f"{dt.isocalendar()[0]}-W{dt.isocalendar()[1]:02d}"
        else:
            key = dt.strftime("%Y-%m")
        groups[key].append(tx)
    return dict(groups)


def _summarise_group(transactions: list) -> dict:
    """Compute aggregate stats for a list of transactions."""
    debit_types = {"send", "withdraw", "paybill", "till", "airtime"}
    credit_types = {"receive", "deposit"}

    total_in = sum(t["amount"] for t in transactions if t.get("type") in credit_types)
    total_out = sum(t["amount"] for t in transactions if t.get("type") in debit_types)
    total_fees = sum(
        t.get("transaction_cost") or 0 for t in transactions if t.get("transaction_cost")
    )

    by_category: Dict[str, float] = defaultdict(float)
    for t in transactions:
        cat = t.get("category") or "Uncategorized"
        if t.get("type") in debit_types:
            by_category[cat] += t["amount"]

    amounts = [t["amount"] for t in transactions]
    avg_tx = statistics.mean(amounts) if amounts else 0.0
    timestamps = [t.get("event_time_utc") or t.get("timestamp") for t in transactions]
    timestamps = [t for t in timestamps if t]
    period_start = min(timestamps) if timestamps else None
    period_end = max(timestamps) if timestamps else None

    return {
        "total_in": total_in,
        "total_out": total_out,
        "net": total_in - total_out,
        "total_fees": total_fees,
        "transaction_count": len(transactions),
        "average_transaction": avg_tx,
        "spending_by_category": dict(by_category),
        "period_start_utc": period_start,
        "period_end_utc": period_end,
        "render_tz": "Africa/Nairobi",
    }


# ─── Public API ───────────────────────────────────────────────────────────────

def weekly_summary(
    db_path: str = "pesa_logger.db",
    weeks: int = 4,
) -> Dict[str, dict]:
    """Return weekly financial summaries for the last *weeks* weeks."""
    now_utc = datetime.now(timezone.utc).replace(tzinfo=None)
    since = now_utc - timedelta(weeks=weeks)
    transactions = list_transactions(db_path=db_path, since=since)
    groups = _group_by_period(transactions, "weekly")
    summary = {k: _summarise_group(v) for k, v in sorted(groups.items())}
    log_report_run(
        report_type="weekly_summary",
        db_path=db_path,
        period_start_utc=since.isoformat(),
        period_end_utc=now_utc.isoformat(),
        tz="Africa/Nairobi",
        output_path=None,
    )
    return summary


def monthly_summary(
    db_path: str = "pesa_logger.db",
    months: int = 6,
) -> Dict[str, dict]:
    """Return monthly financial summaries for the last *months* months."""
    now_utc = datetime.now(timezone.utc).replace(tzinfo=None)
    since = now_utc - timedelta(days=months * 30)
    transactions = list_transactions(db_path=db_path, since=since)
    groups = _group_by_period(transactions, "monthly")
    summary = {k: _summarise_group(v) for k, v in sorted(groups.items())}
    log_report_run(
        report_type="monthly_summary",
        db_path=db_path,
        period_start_utc=since.isoformat(),
        period_end_utc=now_utc.isoformat(),
        tz="Africa/Nairobi",
        output_path=None,
    )
    return summary


def export_csv(
    db_path: str = "pesa_logger.db",
    output_path: Optional[str] = None,
    tx_type: Optional[str] = None,
    since: Optional[datetime] = None,
    until: Optional[datetime] = None,
) -> str:
    """Export transactions to CSV.

    If *output_path* is given the file is written there; otherwise the CSV
    content is returned as a string.
    """
    transactions = list_transactions(
        db_path=db_path, tx_type=tx_type, since=since, until=until
    )
    transactions = sorted(transactions, key=lambda t: t.get("timestamp") or "")
    generated_at = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()

    fieldnames = [
        "transaction_id", "type", "amount", "currency",
        "counterparty_name", "counterparty_phone", "account_number",
        "balance", "running_balance", "transaction_cost", "timestamp",
        "category", "tags", "parser_version", "export_generated_at_utc",
        "render_tz",
    ]
    debit_types = {"send", "withdraw", "paybill", "till", "airtime"}
    credit_types = {"receive", "deposit"}

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    running_balance = 0.0
    rows: List[dict] = []
    for tx in transactions:
        row = dict(tx)
        amount = float(row.get("amount") or 0.0)
        explicit_balance = row.get("balance")
        if explicit_balance is not None:
            running_balance = float(explicit_balance)
        elif row.get("type") in credit_types:
            running_balance += amount
        elif row.get("type") in debit_types:
            running_balance -= amount
        row["running_balance"] = round(running_balance, 2)
        row["export_generated_at_utc"] = generated_at
        row["render_tz"] = "Africa/Nairobi"
        rows.append(row)
    writer.writerows(rows)
    content = buf.getvalue()

    if output_path:
        with open(output_path, "w", newline="", encoding="utf-8") as fh:
            fh.write(content)
    log_report_run(
        report_type="export_csv",
        db_path=db_path,
        period_start_utc=None,
        period_end_utc=None,
        tz="Africa/Nairobi",
        output_path=output_path,
    )

    return content


def export_excel(
    db_path: str = "pesa_logger.db",
    output_path: str = "transactions.xlsx",
    tx_type: Optional[str] = None,
    since: Optional[datetime] = None,
    until: Optional[datetime] = None,
) -> str:
    """Export transactions to an Excel workbook and return the file path."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        ) from exc

    transactions = list_transactions(
        db_path=db_path, tx_type=tx_type, since=since, until=until
    )
    transactions = sorted(transactions, key=lambda t: t.get("timestamp") or "")
    generated_at = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()

    headers = [
        "Transaction ID", "Type", "Amount (KES)", "Counterparty",
        "Phone", "Account", "Balance", "Running Balance", "Fee", "Timestamp",
        "Category", "Tags",
    ]
    field_map = [
        "transaction_id", "type", "amount", "counterparty_name",
        "counterparty_phone", "account_number", "balance", "running_balance",
        "transaction_cost", "timestamp", "category", "tags",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F6B2E")  # Safaricom green

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    debit_types = {"send", "withdraw", "paybill", "till", "airtime"}
    credit_types = {"receive", "deposit"}
    running_balance = 0.0

    for row_idx, tx in enumerate(transactions, start=2):
        row = dict(tx)
        amount = float(row.get("amount") or 0.0)
        explicit_balance = row.get("balance")
        if explicit_balance is not None:
            running_balance = float(explicit_balance)
        elif row.get("type") in credit_types:
            running_balance += amount
        elif row.get("type") in debit_types:
            running_balance -= amount
        row["running_balance"] = round(running_balance, 2)

        for col_idx, field in enumerate(field_map, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field))

    meta = wb.create_sheet("Metadata")
    meta.cell(row=1, column=1, value="generated_at_utc")
    meta.cell(row=1, column=2, value=generated_at)
    meta.cell(row=2, column=1, value="render_tz")
    meta.cell(row=2, column=2, value="Africa/Nairobi")
    meta.cell(row=3, column=1, value="row_count")
    meta.cell(row=3, column=2, value=len(transactions))

    wb.save(output_path)
    log_report_run(
        report_type="export_excel",
        db_path=db_path,
        period_start_utc=None,
        period_end_utc=None,
        tz="Africa/Nairobi",
        output_path=output_path,
    )
    return output_path
